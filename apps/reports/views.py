# apps/reports/views.py
import io
from datetime import date, timedelta

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum, Count, Avg, Q
from django.http import HttpResponse
from django.views.generic import TemplateView, View
from django.utils.timezone import now

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from weasyprint import HTML
from django.template.loader import render_to_string

from apps.fuel.models import FuelRecord
from apps.trips.models import Trip
from apps.maintenance.models import Maintenance


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _date_range_from_request(request):
    """Retorna (date_from, date_to) a partir dos GET params, com fallback de 30 dias."""
    today = date.today()
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    try:
        date_from = date.fromisoformat(date_from_str) if date_from_str else today - timedelta(days=30)
        date_to = date.fromisoformat(date_to_str) if date_to_str else today
    except ValueError:
        date_from = today - timedelta(days=30)
        date_to = today
    return date_from, date_to


def _excel_header_style():
    fill = PatternFill(start_color='01696F', end_color='01696F', fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    align = Alignment(horizontal='center', vertical='center')
    return fill, font, align


def _apply_header(ws, headers):
    fill, font, align = _excel_header_style()
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


# ─────────────────────────────────────────────────────────────
# Painel de Relatórios
# ─────────────────────────────────────────────────────────────

class ReportsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = date.today()
        month_start = today.replace(day=1)

        ctx['fuel_month_total'] = (
            FuelRecord.objects.filter(date__gte=month_start)
            .aggregate(total=Sum('total_cost'))['total'] or 0
        )
        ctx['trips_month_count'] = Trip.objects.filter(start_time__date__gte=month_start).count()
        ctx['maintenance_month_cost'] = (
            Maintenance.objects.filter(date__gte=month_start)
            .aggregate(total=Sum('cost'))['total'] or 0
        )
        ctx['today'] = today
        return ctx


# ─────────────────────────────────────────────────────────────
# COMBUSTÍVEL
# ─────────────────────────────────────────────────────────────

class FuelReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/fuel_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from, date_to = _date_range_from_request(self.request)
        vehicle_id = self.request.GET.get('vehicle')

        qs = FuelRecord.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        agg = qs.aggregate(
            total_cost=Sum('total_cost'),
            total_liters=Sum('liters'),
            avg_price=Avg('price_per_liter'),
            count=Count('id'),
        )

        ctx.update({
            'records': qs.order_by('-date'),
            'date_from': date_from,
            'date_to': date_to,
            'total_cost': agg['total_cost'] or 0,
            'total_liters': agg['total_liters'] or 0,
            'avg_price': agg['avg_price'] or 0,
            'count': agg['count'],
            'vehicle_id': vehicle_id,
        })
        return ctx


class FuelReportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        vehicle_id = request.GET.get('vehicle')

        qs = FuelRecord.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        agg = qs.aggregate(
            total_cost=Sum('total_cost'),
            total_liters=Sum('liters'),
            avg_price=Avg('price_per_liter'),
            count=Count('id'),
        )

        html_string = render_to_string('reports/pdf/fuel_pdf.html', {
            'records': qs.order_by('-date'),
            'date_from': date_from,
            'date_to': date_to,
            'total_cost': agg['total_cost'] or 0,
            'total_liters': agg['total_liters'] or 0,
            'avg_price': agg['avg_price'] or 0,
            'count': agg['count'],
            'generated_at': now(),
        })
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'relatorio_combustivel_{date_from}_{date_to}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class FuelReportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        vehicle_id = request.GET.get('vehicle')

        qs = FuelRecord.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Combustível'

        headers = ['Data', 'Veículo', 'Placa', 'Litros', 'Preço/L (R$)', 'Total (R$)', 'Km', 'Posto']
        _apply_header(ws, headers)

        for r in qs.order_by('-date'):
            ws.append([
                r.date.strftime('%d/%m/%Y'),
                str(r.vehicle),
                r.vehicle.plate,
                float(r.liters),
                float(r.price_per_liter),
                float(r.total_cost),
                r.odometer,
                r.gas_station,
            ])

        # Totais
        ws.append([])
        total_row = ws.max_row + 1
        ws.append(['', '', 'TOTAL', '', '', f'=SUM(F2:F{ws.max_row - 1})', '', ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'relatorio_combustivel_{date_from}_{date_to}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────────────────────
# VIAGENS
# ─────────────────────────────────────────────────────────────

class TripReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/trip_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from, date_to = _date_range_from_request(self.request)
        driver_id = self.request.GET.get('driver')
        vehicle_id = self.request.GET.get('vehicle')

        qs = Trip.objects.filter(
            start_time__date__range=(date_from, date_to)
        ).select_related('vehicle', 'driver', 'driver__user')

        if driver_id:
            qs = qs.filter(driver_id=driver_id)
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        total_distance = sum(t.distance() for t in qs)
        count = qs.count()

        ctx.update({
            'trips': qs.order_by('-start_time'),
            'date_from': date_from,
            'date_to': date_to,
            'total_distance': total_distance,
            'count': count,
            'driver_id': driver_id,
            'vehicle_id': vehicle_id,
        })
        return ctx


class TripReportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        driver_id = request.GET.get('driver')
        vehicle_id = request.GET.get('vehicle')

        qs = Trip.objects.filter(
            start_time__date__range=(date_from, date_to)
        ).select_related('vehicle', 'driver', 'driver__user')
        if driver_id:
            qs = qs.filter(driver_id=driver_id)
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        total_distance = sum(t.distance() for t in qs)

        html_string = render_to_string('reports/pdf/trips_pdf.html', {
            'trips': qs.order_by('-start_time'),
            'date_from': date_from,
            'date_to': date_to,
            'total_distance': total_distance,
            'count': qs.count(),
            'generated_at': now(),
        })
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'relatorio_viagens_{date_from}_{date_to}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class TripReportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        driver_id = request.GET.get('driver')
        vehicle_id = request.GET.get('vehicle')

        qs = Trip.objects.filter(
            start_time__date__range=(date_from, date_to)
        ).select_related('vehicle', 'driver', 'driver__user')
        if driver_id:
            qs = qs.filter(driver_id=driver_id)
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Viagens'

        headers = ['Início', 'Fim', 'Veículo', 'Placa', 'Motorista', 'Destino', 'Propósito', 'Km Inicial', 'Km Final', 'Distância (km)', 'Nº OS']
        _apply_header(ws, headers)

        for t in qs.order_by('-start_time'):
            ws.append([
                t.start_time.strftime('%d/%m/%Y %H:%M'),
                t.end_time.strftime('%d/%m/%Y %H:%M') if t.end_time else '—',
                str(t.vehicle),
                t.vehicle.plate,
                str(t.driver),
                t.destination,
                t.purpose,
                t.start_odometer,
                t.end_odometer or '—',
                t.distance(),
                t.service_order or '—',
            ])

        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'relatorio_viagens_{date_from}_{date_to}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────────────────────
# MANUTENÇÃO
# ─────────────────────────────────────────────────────────────

class MaintenanceReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/maintenance_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from, date_to = _date_range_from_request(self.request)
        vehicle_id = self.request.GET.get('vehicle')
        mtype = self.request.GET.get('type')

        qs = Maintenance.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)
        if mtype:
            qs = qs.filter(type=mtype)

        agg = qs.aggregate(total_cost=Sum('cost'), count=Count('id'))

        ctx.update({
            'records': qs.order_by('-date'),
            'date_from': date_from,
            'date_to': date_to,
            'total_cost': agg['total_cost'] or 0,
            'count': agg['count'],
            'vehicle_id': vehicle_id,
            'selected_type': mtype,
            'maintenance_types': Maintenance.MAINTENANCE_TYPES,
        })
        return ctx


class MaintenanceReportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        vehicle_id = request.GET.get('vehicle')
        mtype = request.GET.get('type')

        qs = Maintenance.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)
        if mtype:
            qs = qs.filter(type=mtype)

        agg = qs.aggregate(total_cost=Sum('cost'), count=Count('id'))

        html_string = render_to_string('reports/pdf/maintenance_pdf.html', {
            'records': qs.order_by('-date'),
            'date_from': date_from,
            'date_to': date_to,
            'total_cost': agg['total_cost'] or 0,
            'count': agg['count'],
            'generated_at': now(),
        })
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'relatorio_manutencao_{date_from}_{date_to}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class MaintenanceReportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)
        vehicle_id = request.GET.get('vehicle')
        mtype = request.GET.get('type')

        qs = Maintenance.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)
        if mtype:
            qs = qs.filter(type=mtype)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Manutenções'

        headers = ['Data', 'Veículo', 'Placa', 'Tipo', 'Descrição', 'Oficina', 'Km', 'Custo (R$)', 'Prox. Alerta (km)', 'Prox. Alerta (data)']
        _apply_header(ws, headers)

        for m in qs.order_by('-date'):
            ws.append([
                m.date.strftime('%d/%m/%Y'),
                str(m.vehicle),
                m.vehicle.plate,
                m.get_type_display(),
                m.description,
                m.workshop,
                m.odometer,
                float(m.cost),
                m.next_alert_km or '—',
                m.next_alert_date.strftime('%d/%m/%Y') if m.next_alert_date else '—',
            ])

        ws.append([])
        ws.append(['', '', '', '', '', 'TOTAL', '', f'=SUM(H2:H{ws.max_row - 1})', '', ''])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'relatorio_manutencao_{date_from}_{date_to}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─────────────────────────────────────────────────────────────
# RELATÓRIO GERAL
# ─────────────────────────────────────────────────────────────

class GeneralReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/general_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        date_from, date_to = _date_range_from_request(self.request)

        fuel_qs = FuelRecord.objects.filter(date__range=(date_from, date_to))
        trip_qs = Trip.objects.filter(start_time__date__range=(date_from, date_to))
        maint_qs = Maintenance.objects.filter(date__range=(date_from, date_to))

        fuel_agg = fuel_qs.aggregate(total_cost=Sum('total_cost'), total_liters=Sum('liters'), count=Count('id'))
        trip_count = trip_qs.count()
        total_distance = sum(t.distance() for t in trip_qs)
        maint_agg = maint_qs.aggregate(total_cost=Sum('cost'), count=Count('id'))

        total_cost = (fuel_agg['total_cost'] or 0) + (maint_agg['total_cost'] or 0)

        ctx.update({
            'date_from': date_from,
            'date_to': date_to,
            # Combustível
            'fuel_cost': fuel_agg['total_cost'] or 0,
            'fuel_liters': fuel_agg['total_liters'] or 0,
            'fuel_count': fuel_agg['count'],
            # Viagens
            'trip_count': trip_count,
            'total_distance': total_distance,
            # Manutenção
            'maint_cost': maint_agg['total_cost'] or 0,
            'maint_count': maint_agg['count'],
            # Total geral
            'total_cost': total_cost,
        })
        return ctx


class GeneralReportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)

        fuel_qs = FuelRecord.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        trip_qs = Trip.objects.filter(start_time__date__range=(date_from, date_to)).select_related('vehicle', 'driver', 'driver__user')
        maint_qs = Maintenance.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')

        fuel_agg = fuel_qs.aggregate(total_cost=Sum('total_cost'), total_liters=Sum('liters'), count=Count('id'))
        maint_agg = maint_qs.aggregate(total_cost=Sum('cost'), count=Count('id'))
        total_distance = sum(t.distance() for t in trip_qs)

        html_string = render_to_string('reports/pdf/general_pdf.html', {
            'date_from': date_from,
            'date_to': date_to,
            'fuel_records': fuel_qs.order_by('-date'),
            'trips': trip_qs.order_by('-start_time'),
            'maintenances': maint_qs.order_by('-date'),
            'fuel_cost': fuel_agg['total_cost'] or 0,
            'fuel_liters': fuel_agg['total_liters'] or 0,
            'fuel_count': fuel_agg['count'],
            'trip_count': trip_qs.count(),
            'total_distance': total_distance,
            'maint_cost': maint_agg['total_cost'] or 0,
            'maint_count': maint_agg['count'],
            'total_cost': (fuel_agg['total_cost'] or 0) + (maint_agg['total_cost'] or 0),
            'generated_at': now(),
        })
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f'relatorio_geral_{date_from}_{date_to}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class GeneralReportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        date_from, date_to = _date_range_from_request(request)

        fuel_qs = FuelRecord.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')
        trip_qs = Trip.objects.filter(start_time__date__range=(date_from, date_to)).select_related('vehicle', 'driver', 'driver__user')
        maint_qs = Maintenance.objects.filter(date__range=(date_from, date_to)).select_related('vehicle')

        wb = openpyxl.Workbook()

        # ── Aba Resumo ──
        ws_summary = wb.active
        ws_summary.title = 'Resumo Geral'
        _apply_header(ws_summary, ['Módulo', 'Qtd. Registros', 'Total (R$)', 'Observação'])

        fuel_agg = fuel_qs.aggregate(total_cost=Sum('total_cost'), total_liters=Sum('liters'), count=Count('id'))
        maint_agg = maint_qs.aggregate(total_cost=Sum('cost'), count=Count('id'))
        total_distance = sum(t.distance() for t in trip_qs)

        ws_summary.append(['Combustível', fuel_agg['count'], float(fuel_agg['total_cost'] or 0), f"{float(fuel_agg['total_liters'] or 0):.2f} litros"])
        ws_summary.append(['Viagens', trip_qs.count(), '—', f"{total_distance} km percorridos"])
        ws_summary.append(['Manutenção', maint_agg['count'], float(maint_agg['total_cost'] or 0), ''])
        ws_summary.append([])
        ws_summary.append(['CUSTO TOTAL', '', float((fuel_agg['total_cost'] or 0) + (maint_agg['total_cost'] or 0)), ''])
        for cell in ws_summary[ws_summary.max_row]:
            cell.font = Font(bold=True)

        for col in ws_summary.columns:
            ws_summary.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        # ── Aba Combustível ──
        ws_fuel = wb.create_sheet('Combustível')
        _apply_header(ws_fuel, ['Data', 'Veículo', 'Placa', 'Litros', 'Preço/L (R$)', 'Total (R$)', 'Km', 'Posto'])
        for r in fuel_qs.order_by('-date'):
            ws_fuel.append([r.date.strftime('%d/%m/%Y'), str(r.vehicle), r.vehicle.plate, float(r.liters), float(r.price_per_liter), float(r.total_cost), r.odometer, r.gas_station])
        for col in ws_fuel.columns:
            ws_fuel.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        # ── Aba Viagens ──
        ws_trips = wb.create_sheet('Viagens')
        _apply_header(ws_trips, ['Início', 'Fim', 'Veículo', 'Motorista', 'Destino', 'Propósito', 'Distância (km)', 'Nº OS'])
        for t in trip_qs.order_by('-start_time'):
            ws_trips.append([t.start_time.strftime('%d/%m/%Y %H:%M'), t.end_time.strftime('%d/%m/%Y %H:%M') if t.end_time else '—', str(t.vehicle), str(t.driver), t.destination, t.purpose, t.distance(), t.service_order or '—'])
        for col in ws_trips.columns:
            ws_trips.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        # ── Aba Manutenção ──
        ws_maint = wb.create_sheet('Manutenção')
        _apply_header(ws_maint, ['Data', 'Veículo', 'Tipo', 'Descrição', 'Oficina', 'Km', 'Custo (R$)'])
        for m in maint_qs.order_by('-date'):
            ws_maint.append([m.date.strftime('%d/%m/%Y'), str(m.vehicle), m.get_type_display(), m.description, m.workshop, m.odometer, float(m.cost)])
        for col in ws_maint.columns:
            ws_maint.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'relatorio_geral_{date_from}_{date_to}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
